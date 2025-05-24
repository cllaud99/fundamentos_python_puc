#   Exercício 9 – Exportação Inteligente
# 	•	Gere um relatório Excel com os seguintes critérios:
# 	•	Clientes de países com temperatura < 15°C
# 	•	AQI acima de 100
# 	•	Receita individual > média geral
# 	•	Utilize OpenPyXL e organize em múltiplas abas: Clientes, Temperaturas, Alertas.


import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(
    0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "src"))
)

from api.airvisual_api import get_aqi
from api.weather_api import get_temperature
from db.db_handler import run_query_from_file
from exercicios_resolucoes.exercicio_04 import obtem_aqi_cidades

QUERY_LIMIT = 10


def obtem_clientes() -> pd.DataFrame:
    """
    Executa consulta SQL para buscar clientes com informações de cidade e país.

    Returns:
        pd.DataFrame: Dados dos clientes com suas respectivas localizações.
    """
    sql_path = "src/db/sql/ex_09_export_excel.sql"
    return run_query_from_file(sql_path, QUERY_LIMIT)


def obtem_aqi(df):
    df = obtem_aqi_cidades(df)
    return df


def obtem_temperatura(df):
    df["temperatura_c"] = df["cidade"].apply(lambda cidade: get_temperature(cidade))
    return df


def filtra_dados(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filtra os dados dos clientes com base nos critérios:
    - Temperatura < 15°C
    - AQI > 100
    - Gasto total > média geral

    Args:
        df (pd.DataFrame): Dados com colunas de temperatura, AQI e gasto.

    Returns:
        pd.DataFrame: Dados filtrados.
    """
    media_gasto = df["gasto_total"].mean()
    df_filtrado = df[
        (df["temperatura_c"] < 15)
        & (df["aqi"] > 100)
        & (df["gasto_total"] > media_gasto)
    ]
    return df_filtrado


def salva_em_excel(
    df_completo: pd.DataFrame, df_filtrado: pd.DataFrame, caminho_arquivo: str
) -> None:
    """
    Cria um arquivo Excel com múltiplas abas:
    - Clientes (completo)
    - Temperaturas
    - Alertas (clientes filtrados)

    Args:
        df_completo (pd.DataFrame): Dados originais dos clientes.
        df_filtrado (pd.DataFrame): Dados filtrados pelos critérios.
        caminho_arquivo (str): Caminho de saída do arquivo Excel.
    """
    wb = Workbook()

    # Aba 1: Clientes
    ws1 = wb.active
    ws1.title = "Clientes"
    for row in dataframe_to_rows(df_completo, index=False, header=True):
        ws1.append(row)

    # Aba 2: Temperaturas
    ws2 = wb.create_sheet(title="Temperaturas")
    colunas_temp = ["nome_completo", "cidade", "pais", "temperatura_c"]
    for row in dataframe_to_rows(df_completo[colunas_temp], index=False, header=True):
        ws2.append(row)

    # Aba 3: Alertas
    ws3 = wb.create_sheet(title="Alertas")
    for row in dataframe_to_rows(df_filtrado, index=False, header=True):
        ws3.append(row)

    wb.save(caminho_arquivo)


def main():
    """
    Função principal para executar o processo de geração do relatório inteligente.
    """
    print("🔄 Buscando dados dos clientes...")
    df = obtem_clientes()
    print(df.head())

    print("🌡️ Buscando temperaturas...")
    df = obtem_temperatura(df)
    print(df.head())

    print("🌫️ Buscando AQI...")
    df = obtem_aqi(df)
    print(df.head())

    print("✅ Filtrando dados...")
    df_filtrado = filtra_dados(df)
    print(df_filtrado.head())

    print("💾 Salvando relatório em Excel...")
    salva_em_excel(df, df_filtrado, "data/reports/relatorio_clientes.xlsx")

    print("🎉 Relatório gerado com sucesso!")


if __name__ == "__main__":
    main()
